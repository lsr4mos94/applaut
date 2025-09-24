from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings 
from .forms import LoginForm, NovoCadastroForm, BonificacaoForm, ItemBonificacaoFormSet
from .models import Cadastro, AnexoCadastro, TotvsCliente, TotvsVendedor, TotvsProduto, TotvsTabPreco, Bonificacao
from django.contrib import messages
from django.views.decorators.http import require_http_methods 
from django.urls import reverse 
from django.contrib.auth import views as auth_views
from django.contrib.auth.views import PasswordResetConfirmView
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db.models import Sum
from django.contrib.auth.models import User
from datetime import datetime
import os
import mimetypes
import openpyxl

@login_required
def inicio(request):
    return render(request, 'inicio.html')

def form(request):
    return render(request, 'form.html')

@login_required
def cadastros(request):
    is_gestor = request.user.groups.filter(name='Gestores').exists()

    if is_gestor:
        cadastros_query = Cadastro.objects.all()
        vendedores_disponiveis = User.objects.filter(groups__name='Vendedores').order_by('first_name', 'username')
    else:
        cadastros_query = Cadastro.objects.filter(vendedor=request.user)
        vendedores_disponiveis = None

    termo_busca = request.GET.get('busca', '').strip()
    if termo_busca:
        cadastros_query = cadastros_query.filter(
            Q(razao_social__icontains=termo_busca) |
            Q(cgc__icontains=termo_busca)
        )

    if is_gestor:
        vendedor_id = request.GET.get('vendedor', '')
        if vendedor_id:
            try:
                vendedor_id = int(vendedor_id)
                cadastros_query = cadastros_query.filter(vendedor_id=vendedor_id)
            except ValueError:
                pass

    situacao = request.GET.get('situacao', '')
    if situacao:
        # Map user-friendly display to actual model status values
        status_map_reverse = {
            'Pendente (Cadastro)': 'pendente',
            'Pendente (Financeiro)': 'cadastrado',
            'Aprovado': 'liberado',
            'Rejeitado': 'rejeitado',
        }
        model_status = status_map_reverse.get(situacao)
        if model_status:
            cadastros_query = cadastros_query.filter(status=model_status) # Changed from situacao_status to status

    cadastros_query = cadastros_query.order_by('-data_cadastro')

    context = {
        'cadastros': cadastros_query,
        'vendedores_disponiveis': vendedores_disponiveis,
        'situacoes_disponiveis': [
            ('Pendente (Cadastro)', 'Pendente (Cadastro)'),
            ('Pendente (Financeiro)', 'Pendente (Financeiro)'),
            ('Aprovado', 'Aprovado'),
            ('Rejeitado', 'Rejeitado'),
        ],
        'termo_busca_selecionado': termo_busca,
        'vendedor_selecionado': request.GET.get('vendedor', ''),
        'situacao_selecionada': situacao,
        'is_gestor': is_gestor,
    }
    return render(request, 'cadastros.html', context)

@login_required
def novo_cadastro(request):
    return render(request, 'novo_cadastro.html')

def novo_cadastro_submit(request):

    if request.method == 'POST':
        form = NovoCadastroForm(request.POST, request.FILES)
        if form.is_valid():
            cadastro = form.save(commit=False)
            
            cadastro.status = 'pendente' 
            
            cadastro.vendedor = request.user 
            cadastro.save()

            anexo_mapping = {
                'cartaoCnpj': 'Cartão CNPJ',
                'contratoSocial': 'Contrato Social/Estatuto',
                'comprovanteEndereco': 'Comprovante de Endereço',
                'documentoSocio': 'Documento de Identificação do Sócio'
            }

            anexos_para_email_attachment = []
            anexos_para_email_context = []

            for input_name, description in anexo_mapping.items():
                if input_name in request.FILES:
                    uploaded_file = request.FILES[input_name]
                    anexo = AnexoCadastro.objects.create(
                        cadastro=cadastro,
                        arquivo=uploaded_file,
                        descricao=description
                    )
                    anexos_para_email_attachment.append(anexo)
                    
                    if anexo.arquivo:
                        filename = os.path.basename(anexo.arquivo.name)
                        anexos_para_email_context.append({'descricao': description, 'nome_arquivo': filename})
                    else:
                        anexos_para_email_context.append({'descricao': description, 'nome_arquivo': 'N/A'})

            confirm_url = request.build_absolute_uri(
                reverse('salesapp:cadastro_confirmar', args=[cadastro.id])
            )
            reject_url = request.build_absolute_uri(
                reverse('salesapp:cadastro_rejeitar', args=[cadastro.id])
            )

            email_context = {
                'cadastro': cadastro,
                'confirm_url': confirm_url,
                'reject_url': reject_url,
                'anexos_info': anexos_para_email_context
            }

            email_html_content = render_to_string('email_cadastro.html', email_context)

            to_email_list = ['cadastroclientes@lautbeer.com.br']

            bcc_email_list = ['lorrane.ramos@lautbeer.com.br']
           
            from_email = settings.DEFAULT_FROM_EMAIL
            subject = f'NOVO CLIENTE PENDENTE DE CADASTRO: #{cadastro.razao_social}'

            email = EmailMessage(
                subject,
                email_html_content,
                from_email,
                to_email_list,
                bcc=bcc_email_list
            )
            email.content_subtype = "html"
            
            for anexo_obj in anexos_para_email_attachment:
                try:
                    if anexo_obj.arquivo and os.path.exists(anexo_obj.arquivo.path):
                        with open(anexo_obj.arquivo.path, 'rb') as f:
                            filename = os.path.basename(anexo_obj.arquivo.name)
                            
                            content_type = None
                            if hasattr(anexo_obj.arquivo, 'file') and hasattr(anexo_obj.arquivo.file, 'content_type'):
                                content_type = anexo_obj.arquivo.file.content_type
                            elif hasattr(anexo_obj.arquivo, 'content_type'):
                                content_type = anexo_obj.arquivo.content_type
                            else:
                                content_type, _ = mimetypes.guess_type(filename)
                            
                            if not content_type:
                                content_type = 'application/octet-stream'
                            
                            email.attach(filename, f.read(), content_type)
                    else:
                        print(f"ATENÇÃO: Arquivo físico não encontrado ou anexo.arquivo é None para {anexo_obj.descricao}. Caminho: {anexo_obj.arquivo.path if anexo_obj.arquivo else 'N/A'}")
                except FileNotFoundError:
                    print(f"ATENÇÃO: Arquivo de anexo não encontrado no caminho: {anexo_obj.arquivo.path}")
                except Exception as e:
                    print(f"ERRO: Falha ao anexar arquivo {anexo_obj.arquivo.name}: {e}")

            try:
                email.send()
            except Exception as e:
                print(f"ERRO: Falha ao enviar e-mail 'email_cadastro.html': {e}")
                messages.error(request, "Erro ao enviar e-mail de novo cadastro. Verifique os logs.")


            return JsonResponse({'success': True, 'message': 'Cadastro enviado e e-mail disparado para aprovação.'})
        else:
            errors_dict = dict(form.errors)
            return JsonResponse({'success': False, 'errors': errors_dict}, status=400)
    else:
        form = NovoCadastroForm()
    return render(request, 'novo_cadastro.html', {'form': form})

@require_http_methods(["GET", "POST"])
def cadastro_confirmar(request, cadastro_id):
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)

    if cadastro.status != 'pendente': 
        context = {'cadastro': cadastro}
        messages.info(request, f"O cadastro de '{cadastro.razao_social}' já foi processado e não pode ser aprovado novamente.")
        return render(request, 'cadastro_ja_processado.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:cadastro_confirmar', args=[cadastro.id])
        return render(request, 'cadastro_confirmar.html', {'cadastro': cadastro, 'action_url': action_url})

    elif request.method == 'POST':
        obs_cadastro_input = request.POST.get('obs', '').strip()

        cadastro.status = 'cadastrado' 
        cadastro.data_aprovacao_cadastro = timezone.now()
        cadastro.obs_cadastro = obs_cadastro_input if obs_cadastro_input else None
        cadastro.save()

        liberar_url = request.build_absolute_uri(
            reverse('salesapp:liberar_cadastro', args=[cadastro.id])
        )
        bloquear_url = request.build_absolute_uri(
            reverse('salesapp:bloquear_cadastro', args=[cadastro.id])
        )
        
        anexos_db = AnexoCadastro.objects.filter(cadastro=cadastro)

        anexos_para_email_context = []
        for anexo_obj in anexos_db:
            if anexo_obj.arquivo:
                filename = os.path.basename(anexo_obj.arquivo.name)
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': filename})
            else:
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': 'N/A'})

        email_context = {
            'cadastro': cadastro,
            'liberar_url': liberar_url,
            'bloquear_url': bloquear_url,
            'anexos_info': anexos_para_email_context,
        }

        email_html_content = render_to_string('email_financeiro.html', email_context)

        to_email_list = ['cobranca@lautbeer.com.br']

        if cadastro.vendedor:
            to_email_list.append(cadastro.vendedor.email)
        else:
            pass

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'NOVO CADASTRO PENDENTE DE LIBERAÇÃO FINANCEIRA: #{cadastro.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc=bcc_email_list
        )
        email.content_subtype = "html"
        
        for anexo_obj in anexos_db:
            try:
                if anexo_obj.arquivo and os.path.exists(anexo_obj.arquivo.path):
                    with open(anexo_obj.arquivo.path, 'rb') as f:
                        filename = os.path.basename(anexo_obj.arquivo.name)
                        
                        content_type = None
                        if hasattr(anexo_obj.arquivo, 'file') and hasattr(anexo_obj.arquivo.file, 'content_type'):
                            content_type = anexo_obj.arquivo.file.content_type
                        elif hasattr(anexo_obj.arquivo, 'content_type'):
                            content_type = anexo_obj.arquivo.content_type
                        else:
                            content_type, _ = mimetypes.guess_type(filename)
                        
                        if not content_type:
                            content_type = 'application/octet-stream'

                        email.attach(filename, f.read(), content_type)
                else:
                    print(f"ATENÇÃO: Arquivo físico não encontrado ou anexo.arquivo é None para {anexo_obj.descricao}. Caminho: {anexo_obj.arquivo.path if anexo_obj.arquivo else 'N/A'}")
            except FileNotFoundError:
                print(f"ATENÇÃO: Arquivo de anexo não encontrado no caminho: {anexo_obj.arquivo.path}")
            except Exception as e:
                print(f"ERRO: Falha ao anexar arquivo {anexo_obj.arquivo.name}: {e}")

        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail de 'pendente_financeiro': {e}")
            messages.error(request, "Erro ao enviar e-mail de notificação financeira. Verifique os logs.")

        messages.success(request, f"O cadastro de '{cadastro.razao_social}' foi aprovado para análise financeira!")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})

    else:
        messages.error(request, "Ação inválida. Use o formulário de aprovação.")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})

@require_http_methods(["GET", "POST"])
def cadastro_rejeitar(request, cadastro_id):
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)

    if cadastro.status not in ['pendente', 'cadastrado']:
        context = {'cadastro': cadastro}
        messages.info(request, f"O cadastro de '{cadastro.razao_social}' já foi processado e não pode ser rejeitado novamente.")
        return render(request, 'cadastro_ja_processado.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:cadastro_rejeitar', args=[cadastro.id])
        return render(request, 'cadastro_rejeitar.html', {'cadastro': cadastro, 'erro_motivo': False, 'action_url': action_url})
        
    elif request.method == 'POST':
        motivo = request.POST.get('motivo_rejeicao', '').strip()
        if not motivo:
            messages.error(request, "O motivo da rejeição não pode estar vazio.")
            action_url = reverse('salesapp:cadastro_rejeitar', args=[cadastro.id])
            return render(request, 'cadastro_rejeitar.html', {'cadastro': cadastro, 'erro_motivo': True, 'action_url': action_url})

        cadastro.status = 'rejeitado'
        cadastro.motivo_rejeicao = motivo
        cadastro.save()

        anexos_db = AnexoCadastro.objects.filter(cadastro=cadastro)

        anexos_para_email_context = []
        for anexo_obj in anexos_db:
            if anexo_obj.arquivo:
                filename = os.path.basename(anexo_obj.arquivo.name)
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': filename})
            else:
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': 'N/A'})

        email_context = {
            'cadastro': cadastro,
            'motivo': motivo,
            'rejeitor_email': request.user.email if request.user.is_authenticated else "rejeitor_desconhecido",
            'anexos_info': anexos_para_email_context,
        }
        email_html_content = render_to_string('cadastro_rejeitado.html', email_context) 

        to_email_list = ['cadastroclientes@lautbeer.com.br']

        if cadastro.vendedor:
            to_email_list.append(cadastro.vendedor.email)
        else:
            pass

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'CLIENTE NÃO CADASTRADO: #{cadastro.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc_email_list
        )
        email.content_subtype = "html"
        
        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail de 'rejeitado_cadastro': {e}")
            messages.error(request, "Erro ao enviar e-mail de notificação de rejeição. Verifique os logs.")

        messages.success(request, f"Cadastro de {cadastro.razao_social} rejeitado na etapa de cadastro. Motivo: {motivo}")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})
    else:
        messages.error(request, "Ação inválida.")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})
    
@require_http_methods(["GET", "POST"])
def liberar_cadastro(request, cadastro_id):
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)

    if cadastro.status != 'cadastrado': 
        context = {'cadastro': cadastro}
        messages.info(request, f"O cadastro de '{cadastro.razao_social}' já foi processado financeiramente e não pode ser liberado novamente.")
        return render(request, 'cadastro_ja_processado.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:liberar_cadastro', args=[cadastro.id])
        return render(request, 'cadastro_confirmar.html', {'cadastro': cadastro, 'action_url': action_url})

    elif request.method == 'POST':
        obs_cadastro_input = request.POST.get('obs', '').strip()

        cadastro.status = 'liberado' 
        cadastro.data_aprovacao_financeiro = timezone.now()
        cadastro.obs_financeiro = obs_cadastro_input if obs_cadastro_input else None
        cadastro.save()

        anexos_db = AnexoCadastro.objects.filter(cadastro=cadastro)

        anexos_para_email_context = []
        for anexo_obj in anexos_db:
            if anexo_obj.arquivo:
                filename = os.path.basename(anexo_obj.arquivo.name)
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': filename})
            else:
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': 'N/A'})


        email_context = {
            'cadastro': cadastro,
            'aprovador_email': request.user.email if request.user.is_authenticated else "aprovador_desconhecido",
            'anexos_info': anexos_para_email_context,
        }
        email_html_content = render_to_string('cadastro_aprovado.html', email_context)

        to_email_list = ['cadastroclientes@lautbeer.com.br', 'cobranca@lautbeer.com.br']

        if cadastro.vendedor:
            to_email_list.append(cadastro.vendedor.email)
        else:
            pass

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'CLIENTE CADASTRADO: #{cadastro.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc_email_list
        )

        email.content_subtype = "html"
        
        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail de 'liberado_cadastro': {e}")
            messages.error(request, "Erro ao enviar e-mail de liberação de cadastro. Verifique os logs.")

        messages.success(request, f"O cadastro de '{cadastro.razao_social}' foi liberado para faturamento.")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})

    else:
        messages.error(request, "Ação inválida.")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})

@require_http_methods(["GET", "POST"])
def bloquear_cadastro(request, cadastro_id):
    cadastro = get_object_or_404(Cadastro, id=cadastro_id)

    if cadastro.status != 'cadastrado': 
        context = {'cadastro': cadastro}
        messages.info(request, f"O cadastro de '{cadastro.razao_social}' já foi processado financeiramente e não pode ser bloqueado novamente.")
        return render(request, 'cadastro_ja_processado.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:bloquear_cadastro', args=[cadastro.id])
        return render(request, 'cadastro_rejeitar.html', {'cadastro': cadastro, 'erro_motivo': False, 'action_url': action_url})

    elif request.method == 'POST':
        motivo = request.POST.get('motivo_rejeicao', '').strip()
        if not motivo:
            messages.error(request, "O motivo da rejeição não pode estar vazio.")
            action_url = reverse('salesapp:bloquear_cadastro', args=[cadastro.id])
            return render(request, 'cadastro_rejeitar.html', {'cadastro': cadastro, 'erro_motivo': True, 'action_url': action_url})

        cadastro.status = 'rejeitado'
        cadastro.motivo_rejeicao = motivo
        cadastro.save()

        anexos_db = AnexoCadastro.objects.filter(cadastro=cadastro)

        anexos_para_email_context = []
        for anexo_obj in anexos_db:
            if anexo_obj.arquivo:
                filename = os.path.basename(anexo_obj.arquivo.name)
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': filename})
            else:
                anexos_para_email_context.append({'descricao': anexo_obj.descricao, 'nome_arquivo': 'N/A'})

        email_context = {
            'cadastro': cadastro,
            'motivo': motivo,
            'rejeitor_email': request.user.email if request.user.is_authenticated else "rejeitor_desconhecido",
            'anexos_info': anexos_para_email_context,
        }
        email_html_content = render_to_string('cadastro_rejeitado.html', email_context) 

        to_email_list = ['cadastroclientes@lautbeer.com.br', 'cobranca@lautbeer.com.br']

        if cadastro.vendedor:
            to_email_list.append(cadastro.vendedor.email)
        else:
            pass

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'CADASTRO NÃO LIBERADO: #{cadastro.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc_email_list
        )
        
        email.content_subtype = "html"
        
        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail de 'bloquear_cadastro': {e}")
            messages.error(request, "Erro ao enviar e-mail de bloqueio de cadastro. Verifique os logs.")

        messages.success(request, f"O cadastro de {cadastro.razao_social} rejeitado pelo financeiro. Motivo: {motivo}")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})
    else:
        messages.error(request, "Ação inválida.")
        return render(request, 'cadastro_ja_processado.html', {'cadastro': cadastro})
    
def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request=request, data=request.POST)

        print("Requisição POST recebida para login.")
        print(f"Dados do formulário: {request.POST}")

        if form.is_valid():
            user = form.get_user()
            print(f"Formulário válido. Usuário autenticado: {user.username}")

            login(request, user)
            print(f"Usuário {user.username} logado com sucesso.")

            return redirect('salesapp:inicio')

        else:
            print("Formulário inválido.")
            print(f"Erros do formulário: {form.errors.as_data()}")
            for field, errors in form.errors.items():
                print(f"Campo '{field}': {errors}")
            print(f"Erros gerais (non_field_errors): {form.non_field_errors()}")

    else:
        form = LoginForm(request=request)
        print("Requisição GET recebida para formulário de login.")

    return render(request, 'login.html', {'form': form})

def user_logout(request):

    logout(request)
    messages.success(request, "Você foi desconectado com sucesso.")
    return redirect('salesapp:login')

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    def form_valid(self, form):
        response = super().form_valid(form)
        return response
    
class CustomPasswordResetView(auth_views.PasswordResetView):
    html_email_template_name = 'registration/password_reset_email.html'

    def post(self, request, *args, **kwargs):

        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)
        return response
    
@login_required
def bonificacoes(request):

    is_gestor = request.user.groups.filter(name='Gestores').exists()

    if is_gestor:
        bonificacao_query = Bonificacao.objects.all()
        vendedores_disponiveis = User.objects.filter(groups__name='Vendedores').order_by('first_name', 'username')
    else:
        bonificacao_query = Bonificacao.objects.filter(vendedor=request.user)
        vendedores_disponiveis = None

    termo_busca = request.GET.get('busca', '').strip()
    if termo_busca:
        bonificacao_query = bonificacao_query.filter(
            Q(razao_social__icontains=termo_busca) |
            Q(cgc__icontains=termo_busca)
        )

    if is_gestor:
        vendedor_id = request.GET.get('vendedor', '')
        if vendedor_id:
            try:
                vendedor_id = int(vendedor_id)
                bonificacao_query = bonificacao_query.filter(vendedor_id=vendedor_id)
            except ValueError:
                pass

    situacao = request.GET.get('situacao', '')
    if situacao:
        bonificacao_query = bonificacao_query.filter(status=situacao)

    data_de_str = request.GET.get('data_inicio', '')
    data_ate_str = request.GET.get('data_fim', '')

    if data_de_str:
        try:
            data_inicio = datetime.strptime(data_de_str, '%Y-%m-%d').date()
            bonificacao_query = bonificacao_query.filter(data_criacao__date__gte=data_inicio)
        except ValueError:
            pass
            
    if data_ate_str:
        try:
            data_fim = datetime.strptime(data_ate_str, '%Y-%m-%d').date()
            bonificacao_query = bonificacao_query.filter(data_criacao__date__lte=data_fim)
        except ValueError:
            pass

    bonificacao_query = bonificacao_query.order_by('-data_criacao')

    context = {
        'bonificacoes': bonificacao_query,
        'vendedores_disponiveis': vendedores_disponiveis,
        'situacoes_disponiveis': [
            ('PENDENTE_GESTOR', 'Pendente Aprovação'),
            ('PENDENTE_PEDIDO', 'Pendente Pedido'),
            ('RECUSADA', 'Recusada'),
            ('PEDIDO_GERADO', 'Pedido Gerado'),
        ],
        'termo_busca_selecionado': termo_busca,
        'vendedor_selecionado': request.GET.get('vendedor', ''),
        'situacao_selecionada': situacao,
        'is_gestor': is_gestor,
    }
    return render(request, 'bonificacoes.html', context)

@login_required
def nova_bonificacao(request):
    if request.method == 'POST':
        pass
    else: 
        bonificacao_form = BonificacaoForm()
        item_formset = ItemBonificacaoFormSet(prefix='itens')

    context = {
        'form': bonificacao_form,
        'formset': item_formset,
    }
    return render(request, 'nova_bonificacao.html', context)

@require_POST
def nova_bonificacao_submit(request):
    if request.method == 'POST':
        form = BonificacaoForm(request.POST, request.FILES)
        
        if form.is_valid():
            bonificacao = form.save(commit=False)
            bonificacao.vendedor = request.user 
            
            bonificacao.save()

            try:
                item_formset = ItemBonificacaoFormSet(request.POST, request.FILES, instance=bonificacao, prefix='itens') 
            except Exception as e:
                print(f"ERRO: Falha ao instanciar ItemBonificacaoFormSet: {e}")
                return JsonResponse(
                    {'success': False, 'errors': {'formset_creation': f'Erro ao processar itens do formulário: {str(e)}'}},
                    status=400
                )

            if item_formset.is_valid():

                item_formset.save() 
                
                tipo_bonificacao = form.cleaned_data['tipo_bonificacao']

                if tipo_bonificacao in ['VERBA_CONTRATO', 'BONIFICACAO_VERBA']:
                    bonificacao.status = 'PEDIDO_GERADO' 
                    total_bonificacao = bonificacao.itens.aggregate(total=Sum('valor_total_item'))['total'] or 0.00
                    bonificacao.valor_total = total_bonificacao
                    bonificacao.save(update_fields=['status', 'valor_total'])

                    return JsonResponse({'success': True, 'message': 'Solicitação registrada e autorizada. Favor inserir o pedido no sistema.'})
                
                elif tipo_bonificacao == 'NEGOCIACAO_ESPECIAL':
                    bonificacao.status = 'PENDENTE_GESTOR' 
                    total_bonificacao = bonificacao.itens.aggregate(total=Sum('valor_total_item'))['total'] or 0.00
                    bonificacao.valor_total = total_bonificacao
                    bonificacao.save(update_fields=['status', 'valor_total'])

                    confirm_url = request.build_absolute_uri(
                        reverse('salesapp:bonificacao_autorizar_gestor', args=[bonificacao.id])
                    )
                    reject_url = request.build_absolute_uri(
                        reverse('salesapp:bonificacao_recusar_gestor', args=[bonificacao.id])
                    )

                    email_context = {
                        'bonificacao': bonificacao,
                        'itens': bonificacao.itens.all(), 
                        'confirm_url': confirm_url,
                        'reject_url': reject_url,
                    }

                    email_html_content = render_to_string('email_gestor.html', email_context)
                    to_email_list = ['wellington@lautbeer.com.br']
                    bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

                    from_email = settings.DEFAULT_FROM_EMAIL
                    subject = f'BONIFICAÇÃO PENDENTE DE AUTORIZAÇÃO: {bonificacao.razao_social}'

                    email = EmailMessage(
                        subject,
                        email_html_content,
                        from_email,
                        to_email_list, 
                        bcc=bcc_email_list
                    )
                    email.content_subtype = "html"
                    
                    try:
                        email.send()
                    except Exception as e:
                        print(f"ERRO: Falha ao enviar e-mail 'email_gestor.html': {e}")
                        messages.error(request, "Erro ao enviar e-mail de nova bonificação. Verifique os logs.")

                    return JsonResponse({'success': True, 'message': 'Solicitação enviada e e-mail disparado para aprovação.'})
                else:
                    return JsonResponse({'success': False, 'errors': 'Tipo de bonificação inválido.'}, status=400)
            else:
                print(f"Erros no formset de itens: {item_formset.errors}")
                item_errors_data = []
                for form_errors_for_item in item_formset.errors:
                    if form_errors_for_item:
                        item_error_dict = {}
                        for field, errors_list in form_errors_for_item.items():
                            item_error_dict[field] = [str(err) for err in errors_list]
                        item_errors_data.append(item_error_dict)
                
                return JsonResponse({'success': False, 'errors': {'item_forms': item_errors_data}}, status=400)
        else:
            print(f"Erros no formulário principal: {form.errors}")
            main_form_errors_dict = {}
            for field, errors_list in form.errors.items():
                main_form_errors_dict[field] = [str(err) for err in errors_list]
            
            if form.non_field_errors():
                main_form_errors_dict['non_field_errors'] = [str(error) for error in form.non_field_errors()]

            return JsonResponse({'success': False, 'errors': {'main_form': main_form_errors_dict}}, status=400)
    else: 
        form = BonificacaoForm() 
    
    return render(request, 'nova_bonificacao.html', {'form': form})

@login_required
def buscar_clientes(request):
    termo_busca = request.GET.get('q', '').strip()
    clientes_encontrados = []

    if not termo_busca:
        return JsonResponse({'clientes': []})

    try:
        email_usuario_logado = request.user.email

        vendedor = TotvsVendedor.objects.filter(email_vendedor=email_usuario_logado).first()
        cod_vendedor_logado = vendedor.cod_vendedor

        clientes = TotvsCliente.objects.filter(
            Q(razao_social__icontains=termo_busca) |
            Q(nome_fantasia__icontains=termo_busca) |
            Q(cgc__icontains=termo_busca),
            cod_vendedor=cod_vendedor_logado
        ).values('razao_social', 'nome_fantasia', 'cgc', 'grp_cliente', 'cod_cliente', 'loja_cliente').distinct()

        for cliente in clientes:
            clientes_encontrados.append({
                'razaoSocial': cliente['razao_social'],
                'nomeFantasia': cliente['nome_fantasia'],
                'cgc': cliente['cgc'],
                'grupoCliente': cliente['grp_cliente'],
                'codigoCliente': cliente['cod_cliente'],
                'lojaCliente': cliente['loja_cliente'],
            })

    except TotvsVendedor.DoesNotExist:
        return JsonResponse({'clientes': [], 'error': 'Vendedor não encontrado para o usuário logado.'}, status=404)
    except Exception as e:
        return JsonResponse({'clientes': [], 'error': str(e)}, status=500)

    return JsonResponse({'clientes': clientes_encontrados})

@login_required
def buscar_produtos(request):
    termo_busca = request.GET.get('q', '').strip()
    produtos_encontrados = []

    if not termo_busca:
        return JsonResponse({'produtos': []})

    try:
        produtos_qs = TotvsProduto.objects.filter(
            Q(cod_produto__icontains=termo_busca) |
            Q(desc_produto__icontains=termo_busca)
        ).values('cod_produto', 'desc_produto').distinct()


        if not produtos_qs.exists():
            return JsonResponse({'produtos': []})

        for produto in produtos_qs:
            cod_produto = produto['cod_produto']
            descricao_produto = produto['desc_produto']
            preco_tabela = "0.00"

            try:
                preco_obj = TotvsTabPreco.objects.filter(pk=cod_produto).first()

                if preco_obj:
                    if hasattr(preco_obj, 'vlr_unitario'):
                        preco_tabela = str(preco_obj.vlr_unitario)
                    elif hasattr(preco_obj, 'preco_unitario'): 
                        preco_tabela = str(preco_obj.preco_unitario)
                    elif hasattr(preco_obj, 'val_unitario'):
                        preco_tabela = str(preco_obj.val_unitario)
                    elif hasattr(preco_obj, 'preco'):
                        preco_tabela = str(preco_obj.preco)
                    else:
                        pass
                else:
                        pass
            except Exception as e_preco:
                pass

            produtos_encontrados.append({
                'codigoProduto': cod_produto,
                'descricaoProduto': descricao_produto,
                'precoTabela': preco_tabela,
            })

        return JsonResponse({'produtos': produtos_encontrados})

    except Exception as e:
        return JsonResponse({'produtos': [], 'error': str(e)}, status=500)

@require_http_methods(["GET", "POST"])
def bonificacao_recusar_gestor(request, bonificacao_id):
    bonificacao = get_object_or_404(Bonificacao, id=bonificacao_id)

    if bonificacao.status != 'PENDENTE_GESTOR':
        context = {'bonificacao': bonificacao}
        messages.info(request, f"A solicitação de bonificação para '{bonificacao.razao_social}' já foi processada e não pode ser alterado.")
        return render(request, 'bonificacao_ja_processada.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:bonificacao_recusar_gestor', args=[bonificacao.id])
        return render(request, 'bonificacao_recusar.html', {'bonificacao': bonificacao, 'erro_motivo': False, 'action_url': action_url})
        
    elif request.method == 'POST':
        motivo = request.POST.get('motivo_rejeicao', '').strip()
        if not motivo:
            messages.error(request, "O motivo não pode estar vazio.")
            action_url = reverse('salesapp:bonificacao_recusar', args=[bonificacao.id])
            return render(request, 'bonificacao_recusar.html', {'bonificacao': bonificacao, 'erro_motivo': True, 'action_url': action_url})

        bonificacao.status = 'RECUSADA'
        bonificacao.motivo_recusa = motivo
        bonificacao.save()

        bonificacao_itens = bonificacao.itens.all()

        email_context = {
            'bonificacao': bonificacao,
            'itens': bonificacao_itens,
            'motivo': motivo,
            'rejeitor_email': request.user.email if request.user.is_authenticated else "rejeitor_desconhecido",
        }

        email_html_content = render_to_string('bonificacao_recusada.html', email_context) 

        to_email_list = ['wellington@lautbeer.com.br']

        if bonificacao.vendedor:
            to_email_list.append(bonificacao.vendedor.email)
        else:
            pass

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'BONIFICAÇÃO NÃO AUTORIZADA: {bonificacao.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc_email_list
        )
        email.content_subtype = "html"
        
        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail de 'bonificacao_recusada': {e}")
            messages.error(request, "Erro ao enviar e-mail de notificação de recusa. Verifique os logs.")

        messages.success(request, f"Solicitação de bonificacaçõ para {bonificacao.razao_social} rejeitada pelo gestor. Motivo: {motivo}")
        return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})
    else:
        messages.error(request, "Ação inválida.")
        return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})
    
@require_http_methods(["GET", "POST"])
def bonificacao_autorizar_gestor(request, bonificacao_id):
    
    bonificacao = get_object_or_404(Bonificacao, id=bonificacao_id)

    if bonificacao.status != 'PENDENTE_GESTOR': 
        context = {'bonificacao': bonificacao}
        messages.info(request, f"A solicitação de bonificação para '{bonificacao.razao_social}' já foi processada e não pode ser alterado.")
        return render(request, 'bonificacao_ja_processada.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:bonificacao_autorizar_gestor', args=[bonificacao.id])
        return render(request, 'bonificacao_autorizar.html', {'bonificacao': bonificacao, 'action_url': action_url})

    elif request.method == 'POST':
        obs_gestor_input = request.POST.get('obs', '').strip()

        bonificacao.status = 'PENDENTE_PEDIDO' 
        bonificacao.data_aprovacao_gestor = timezone.now()
        bonificacao.obs_diretoria = obs_gestor_input if obs_gestor_input else ''
        bonificacao.save()

        confirm3_url = request.build_absolute_uri(
            reverse('salesapp:bonificacao_pedido', args=[bonificacao.id])
        )

        bonificacao_itens = bonificacao.itens.all()

        email_context = {
            'bonificacao': bonificacao,
            'itens': bonificacao_itens,
            'confirm_url': confirm3_url,
        }

        email_html_content = render_to_string('email_pedido.html', email_context)

        to_email_list = ['admcomercial@lautbeer.com.br']

        bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

        from_email = settings.DEFAULT_FROM_EMAIL
        subject = f'BONIFICAÇÃO AUTORIZADA PENDENTE DE EMISSÃO DO PEDIDO: {bonificacao.razao_social}'

        email = EmailMessage(
            subject,
            email_html_content,
            from_email,
            to_email_list,
            bcc=bcc_email_list
        )
        email.content_subtype = "html"
        
        try:
            email.send()
        except Exception as e:
            print(f"ERRO: Falha ao enviar e-mail 'email_pedido.html': {e}")
            messages.error(request, "Erro ao enviar e-mail de nova bonificação. Verifique os logs.")

        messages.success(request, f'Solicitação aprovada e e-mail enviado para emissão do pedido.')
        return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})

    else:
        messages.error(request, "Ação inválida. Use o formulário de aprovação.")
        return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})         
    
@require_http_methods(["GET", "POST"])
def bonificacao_pedido(request, bonificacao_id):
    bonificacao = get_object_or_404(Bonificacao, id=bonificacao_id)

    if bonificacao.status != 'PENDENTE_PEDIDO': 
        context = {'bonificacao': bonificacao}
        messages.info(request, f"A solicitação de bonificação para '{bonificacao.razao_social}' já foi processada e não pode ser alterado.")
        return render(request, 'bonificacao_ja_processada.html', context)

    if request.method == 'GET':
        action_url = reverse('salesapp:bonificacao_pedido', args=[bonificacao.id])
        return render(request, 'bonificacao_pedido.html', {'bonificacao': bonificacao, 'action_url': action_url})

    elif request.method == 'POST':
        status_emissao = request.POST.get('status')
        obs_pedido_input = request.POST.get('obs', '')
        numero_pedido = request.POST.get('numero_pedido', '')

        bonificacao_itens = bonificacao.itens.all()

        if status_emissao == 'emitido':

            bonificacao.status = 'PEDIDO_GERADO'

            bonificacao.data_pedido = timezone.now()
            bonificacao.obs_pedido = obs_pedido_input if obs_pedido_input else ''
            bonificacao.numero_pedido = numero_pedido
            bonificacao.save()

            email_context = {
            'bonificacao': bonificacao,
            'itens': bonificacao_itens,
            }

            email_html_content = render_to_string('bonificacao_autorizada.html', email_context)

            to_email_list = ['admcomercial@lautbeer.com.br', 'wellington@lautbeer.com.br']

            if bonificacao.vendedor:
                to_email_list.append(bonificacao.vendedor.email)
            else:
                pass

            bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

            from_email = settings.DEFAULT_FROM_EMAIL
            subject = f'PEDIDO DE BONIFICAÇÃO EMITIDO: {bonificacao.razao_social}'

            email = EmailMessage(
                subject,
                email_html_content,
                from_email,
                to_email_list,
                bcc=bcc_email_list
            )
            email.content_subtype = "html"
        
            try:
                email.send()
            except Exception as e:
                print(f"ERRO: Falha ao enviar e-mail 'bonificacao_autorizada.html': {e}")
                messages.error(request, "Erro ao enviar e-mail de pedido emitido. Verifique os logs.")

            messages.success(request, f'Solicitação aprovada e pedido emitido.')
            return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})
            
        elif status_emissao == 'nao_emitido':

            bonificacao.status = 'RECUSADA' 
            bonificacao.motivo_recusa = obs_pedido_input if obs_pedido_input else ''
            bonificacao.save()

            email_context = {
            'bonificacao': bonificacao,
            'itens': bonificacao_itens,
            }

            email_html_content = render_to_string('bonificacao_recusada.html', email_context)
            
            to_email_list = ['admcomercial@lautbeer.com.br', 'wellington@lautbeer.com.br']

            if bonificacao.vendedor:
                to_email_list.append(bonificacao.vendedor.email)
            else:
                pass

            bcc_email_list = ['lorrane.ramos@lautbeer.com.br']

            from_email = settings.DEFAULT_FROM_EMAIL
            subject = f'PEDIDO DE BONIFICAÇÃO NÃO EMITIDO: {bonificacao.razao_social}'

            email = EmailMessage(
                subject,
                email_html_content,
                from_email,
                to_email_list,
                bcc=bcc_email_list
            )
            email.content_subtype = "html"
        
            try:
                email.send()
            except Exception as e:
                print(f"ERRO: Falha ao enviar e-mail 'bonificacao_recusada.html': {e}")
                messages.error(request, "Erro ao enviar e-mail de pedido não emitido. Verifique os logs.")

            messages.success(request, f'Pedido de bonificação não emitido.')
            return render(request, 'bonificacao_ja_processada.html', {'bonificacao': bonificacao})
        
def export_bonificacoes_xlsx(request):

    termo_busca = request.GET.get('busca', '')
    vendedor_id = request.GET.get('vendedor', '')
    situacao = request.GET.get('situacao', '')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    bonificacoes = Bonificacao.objects.all()

    if termo_busca:
        bonificacoes = bonificacoes.filter(
            Q(razao_social__icontains=termo_busca) | Q(cgc__icontains=termo_busca)
        )

    if vendedor_id:
        bonificacoes = bonificacoes.filter(vendedor_id=vendedor_id)

    if situacao:
        bonificacoes = bonificacoes.filter(status=situacao)

    if data_inicio_str:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        bonificacoes = bonificacoes.filter(data_criacao__date__gte=data_inicio)

    if data_fim_str:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        bonificacoes = bonificacoes.filter(data_criacao__date__lte=data_fim)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=bonificacoes.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bonificações"

    columns = [
        "Razão Social",
        "CGC",
        "Vendedor",
        "Data da Solicitação",
        "Valor Total",
        "Situação",
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for bonificacao in bonificacoes:
        row_num += 1

        cgc_str = str(bonificacao.cgc)
        status_display = bonificacao.get_status_display()
        data_sem_tz = bonificacao.data_criacao.replace(tzinfo=None)

        worksheet.cell(row=row_num, column=1, value=bonificacao.razao_social)
        worksheet.cell(row=row_num, column=2, value=cgc_str)
        worksheet.cell(row=row_num, column=3, value=bonificacao.vendedor.get_full_name())
        worksheet.cell(row=row_num, column=4, value=data_sem_tz)
        worksheet.cell(row=row_num, column=5, value=bonificacao.valor_total)
        worksheet.cell(row=row_num, column=6, value=status_display)

    workbook.save(response)

    return response